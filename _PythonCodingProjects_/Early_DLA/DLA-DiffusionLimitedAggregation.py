from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import openpyxl
import random
import math

wb = load_workbook("DLA.xlsx")
ws = wb.active

size = 102
dop = .1 #Density Of Particles in solution represented with decimal
iterations = 100


#Initializer
#Sets all vals to 0
for row in range(1,size+1):
    for col in range(1,size+1):
        char = get_column_letter(col)
        ws[char + str(row)] = 0

#Sets particles in solution with dop chance
for row in range(1,size+1):
    for col in range(1,size+1):
        char = get_column_letter(col)
        if random.random() < dop:
            ws[char + str(row)] = 1


#Seed Placer
ws[get_column_letter(round(size/2)) + str(round(size/2))] = 2


#Selects random neighbor to move to and checks if neighbor is occupied.
#If unoccupied, the new location is returned, and if unoccupied, occupied neighbor is removed and loop restarts.
#If there are no avalible neighbors, the original value is returned.
#If the value is unable to be selected (due to being out of bounds), the value is removed and loop restarts. 
def move(cor):
    listGroup = [[-1,-1], [0,-1], [1,-1], [-1,0], [1,0], [-1,1], [0,1], [1,1]]
    #listGroup = [[0, -1], [-1, 0], [1, 0], [0, 1]]
    while (len(listGroup) > 0):
        groupCor = random.choice(listGroup)
        newCor = [cor[i] + groupCor[i] for i in range(len(cor))]
        try:
            if (ws[get_column_letter(newCor[0]) + str(newCor[1])].value != 0):
                listGroup.remove(groupCor)
                continue
            else:
                return newCor
        except ValueError:
            listGroup.remove(groupCor)
            continue
    return cor
      
def goTo(oldCor, newCor):
    oldChar, newChar = get_column_letter(oldCor[0]), get_column_letter(newCor[0])
    ws[oldChar + str(oldCor[1])] = 0
    ws[newChar + str(newCor[1])] = 1

def stick(cor):
    listGroup = [[-1,-1], [0,-1], [1,-1], [-1,0], [1,0], [-1,1], [0,1], [1,1]]
    #listGroup = [[0, -1], [-1, 0], [1, 0], [0, 1]]
    while (len(listGroup) > 0):
        groupCor = listGroup[0]
        newCor = [cor[i] + groupCor[i] for i in range(len(cor))]
        try:
            if (ws[get_column_letter(newCor[0]) + str(newCor[1])].value == 2):
                ws[get_column_letter(cor[0]) + str(cor[1])] = 3
                return
            else:
                listGroup.remove(groupCor)
                continue    
        except ValueError:
            listGroup.remove(groupCor)
    return 
            
    
        
while (iterations != 0):
    
    for row in range(1,size+1):
        for col in range(1,size+1):
            char = get_column_letter(col)
            if ws[char + str(row)].value == 1:
                goTo([col, row], move([col, row]))
                stick([col, row])
                

    for row in range(1,size+1):
        for col in range(1,size+1):
            char = get_column_letter(col)
            if ws[char + str(row)].value == 3:
                ws[char + str(row)] = 2

    iterations -= 5
#''')

wb.save("DLA.xlsx")
print("Finished.")


