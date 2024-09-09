from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import openpyxl
import random
import math

wb = load_workbook("CityTestV1.xlsx")
ws = wb.active

size = 100

#Initializer
#Sets all vals to 1
for row in range(1,size+1):
    for col in range(1,size+1):
        char = get_column_letter(col)
        ws[char + str(row)] = 0

col, row = int(size/2), int(size/2)


def rndwlk(itr, size, col, row, val):
    for i in range(itr):
        
        nxtspot = random.choice([1, -1])
        if random.randint(1,2) == 1:
            for i in range(random.choice([2, 4, 6])):
                col += nxtspot
                ws[get_column_letter(col) + str(row)] = val        
        else:
            if random.randint(1,20) != 1:
                for i in range(random.choice([2, 4, 6])):
                    row += nxtspot
                    ws[get_column_letter(col) + str(row)] = val
            else:
                for i in range(random.choice([2, 4, 6])):
                    row += nxtspot
                    for i in range(random.randint(1,7)):
                        col += nxtspot
                        ws[get_column_letter(col) + str(row)] = val
                    ws[get_column_letter(col) + str(row)] = val
                    
                     
        if (abs(col - int(size/2)) > int(size/4) or abs(row - int(size/2)) > int(size/4)):
            col, row = int(size/2), int(size/2)

for i in range(5):
    rndwlk(random.choice([100, 200, 300]), size, col, row, i)
        
wb.save("CityTestV1.xlsx")
print("Finished.")
