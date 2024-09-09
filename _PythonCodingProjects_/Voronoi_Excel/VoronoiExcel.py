from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils.cell import column_index_from_string
import openpyxl
import random
import math

wb = load_workbook("VoronoiExcel.xlsx")
ws = wb.active

size = 200
keyPts = 30

#Initializer
#Sets all vals to 1
for row in range(1,size+1):
    for col in range(1,size+1):
        char = get_column_letter(col)
        ws[char + str(row)] = 1


ptsArr = []
#Sets keyPts
for i in range(keyPts):
    char = get_column_letter(random.randint(1,size))
    row = str(random.randint(1,size))
    pt = char + row
    ws[pt] = 0
    ptsArr.append(pt)

ptsVal = []
for i in range(len(ptsArr)):
    ptsVal.append(i+1)

for row in range(1,size+1):
    for col in range(1,size+1):
        char = get_column_letter(col)
        
        smlstdist = size * (2**.5) + 1 #Max possible dist plus 1
        closePt = 'A' + '1'
        #Finds which of the keyPts is closest
        for i in range(len(ptsArr)):
            keyCor = coordinate_from_string(ptsArr[i])
            keyCol = column_index_from_string(keyCor[0])
            keyRow = keyCor[1]
            dist = ((abs(keyCol - col) ** 2) + (abs(keyRow - row) ** 2)) ** .5 #Pythagorean theorem
            if dist < smlstdist:
                smlstdist = dist
                closePt = ptsArr[i]

        ws[char + str(row)] = ptsVal[ptsArr.index(closePt)]

for i in range(len(ptsArr)):
    ws[ptsArr[i]] = 0
        
wb.save("VoronoiExcel.xlsx")
print("Finished.")
