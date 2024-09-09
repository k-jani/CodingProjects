from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
import random
import math

wb = load_workbook("T2.xlsx")
ws = wb.active

size = 100

#Initializer
for row in range(1,size+1):
    for col in range(1,size+1):
        char = get_column_letter(col)
        ws[char + str(row)] = random.randint(1,2) #Set val of col,row to distval func of the x,y coords

#Avg func
def avgfunc(col, row):
    if col == 1 or row == 1 or row == size or col == size:
        return ws[get_column_letter(col) + str(row)].value #Returns the same value of the func of cells on border
    else:
        addavg = ws[get_column_letter(col+1) + str(row)].value #Value to the right
        addavg += ws[get_column_letter(col-1) + str(row)].value #Value to the left
        addavg += ws[get_column_letter(col) + str(row-1)].value #Value up
        addavg += ws[get_column_letter(col) + str(row+1)].value #Value down
        finval = addavg / 4
        return finval
    
#Initial averager
for row in range(1,size+1):
    for col in range(1,size+1):
        char = get_column_letter(col)
        ws[char + str(row)] = avgfunc(col, row) #Set val of col,row to distval func of the x,y coords

#Smoother function
for i in range(1,5): #Where second number defines strength of avging
    for row in range(1,size+1):
        for col in range(1,size+1):
            char = get_column_letter(col)
            ws[char + str(row)] = avgfunc(col, row) #Set val of col,row to distval func of the x,y coords
avgttl = 0
for row in range(1,size+1):
    for col in range(1,size+1):
        avgttl += ws[char + str(row)].value

avgttl /= (size ** 2)

#Polarization
for row in range(1,size+1):
    for col in range(1,size+1):
        char = get_column_letter(col)
        if ws[char + str(row)].value > avgttl:
            ws[char + str(row)] = ws[char + str(row)].value + 1
        elif ws[char + str(row)].value < avgttl:
            ws[char + str(row)] = ws[char + str(row)].value - 1
        else:
            ws[char + str(row)] = ws[char + str(row)].value + 0

#Smoother function
for i in range(1,4): #Where second number defines strength of avging
    for row in range(1,size+1):
        for col in range(1,size+1):
            char = get_column_letter(col)
            ws[char + str(row)] = avgfunc(col, row) #Set val of col,row to distval func of the x,y coords            
wb.save("T2.xlsx")
